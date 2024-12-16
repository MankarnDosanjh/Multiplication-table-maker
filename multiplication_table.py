'''multiplication_table.py
Create a program multiplicationTable.py that takes a number N and creates an NxN multiplication table in an Excel spreadsheet.'''

# Sets working directory spreadsheet will be stored in
import os, sys
os.chdir(os.path.dirname(sys.argv[0]))

# Main module imports
from openpyxl import Workbook, utils
from openpyxl.styles import Font, colors

# Makes sure user enters value above 0 that qualifies for a timestable
while True:
    try:
        multiple = int(input('Please enter any number greater than 0:\n'))
        if multiple < 1:
            raise ValueError
        break
    except:
        print('ERROR: Invalid input!\n')

# Initialises workbook and worksheet
wb = Workbook()
sheet = wb.active

# Creates auto colour and bold font objects
auto_colour = colors.Color.auto = True
bold = Font(bold=True)

# Generates table headings
for x in range(2, multiple + 2): # Range accounts for openpyxl starting count from 1
    column_header = sheet.cell(row=x, column=1)
    row_header = sheet.cell(row=1, column=x)

    # Assigns multiple in header, accounting for range offset
    column_header.value = x - 1
    row_header.value = x - 1

    # Makes headers bold
    column_header.font = bold
    row_header.font = bold

# Stores size of table for iteration
max_row = sheet.max_row
max_column = utils.get_column_letter(sheet.max_column)

# Iterates through remaining empty cells
for cell_row in sheet['B2': f'{max_column}{max_row}']:
    for cell in cell_row:
        # Stores multiples in row and column relative to cell
        cur_column = utils.get_column_letter(cell.column)
        row_mult = sheet[f'A{cell.row}'].value
        col_mult = sheet[f'{cur_column}1'].value

        # Assigns multiplication result and font to cell
        cell.value = row_mult * col_mult
        cell.font = auto_colour

wb.save(f'./{multiple}_times_table.xlsx')